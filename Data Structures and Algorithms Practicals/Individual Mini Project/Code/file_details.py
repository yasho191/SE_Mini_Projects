import os
import openpyxl
import string


# get the details about the parameters
def get_details(ws):
    columns = 0
    student_eval_parameters = list()
    # iterate over the columns of the first row until we encounter an empty cell
    for col in ws.iter_cols(max_row=1):
        for cell in col:
            if cell.value is None:
                break
            student_eval_parameters.append(cell.value)
            columns += 1

    return columns, student_eval_parameters


# Read all the records from the file
def get_records(ws, columns):
    student_records = dict()

    for rows in ws.iter_rows(min_row=2):
        i = 0
        record = dict()
        for cell in rows:
            if cell.value is None:
                break

            if columns[i] == "Roll Number":
                roll = cell.value
            else:
                record[columns[i]] = cell.value
            i += 1

        if i == 0:
            break
        student_records[roll] = record

    return student_records


# get the max marks for all parameters used for evaluation.
def get_max_values(columns, sep):
    print("Enter 1 for parameters to be considered for evaluation of marks else enter 0: ")
    print()
    mep = [0 for i in range(columns)]
    j = 0
    for parameter in sep:
        mep[j] = int(input(f"{j + 1}. {parameter}".ljust(50) + ": "))
        j += 1
    print("Data stored successfully!")
    print()
    print("Enter Maximum Marks for the Selected Parameters: ")
    max_values = dict()
    for i in range(columns):
        if mep[i]:
            if sep[i].lower() == "attendance":
                days = int(input("Enter the Total Days: "))
                max_values[sep[i]] = int(input("Enter the Total Marks for attendance: "))
            else:
                max_values[sep[i]] = int(input(f"Enter Maximum Marks for {sep[i]}: "))

    return max_values, days


def put_results(ws, results, columns):
    alphas = list(string.ascii_uppercase)
    cell = f"{alphas[columns - 1]}1"
    ws[cell] = "Results"
    for i in range(2, len(results) + 2):
        cell = f"{alphas[columns - 1]}{i}"
        ws[cell] = results[i - 2]
